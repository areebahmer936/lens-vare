"""
LensVare — Audit Intelligence Backend
======================================
FastAPI service exposing the Excel → SQLite → Text-to-SQL → Prose pipeline
as a REST API consumed by the frontend.

Endpoints
---------
POST /upload            — Upload Excel file; returns session_id + schema summary
POST /ask               — Natural language query; returns prose + sql + row_count
GET  /sessions/{id}/schema  — Return the schema text for a session
DELETE /sessions/{id}   — Drop a session from memory
GET  /providers         — List available providers and their models
GET  /health            — Liveness probe for Docker/hosting health checks
"""

import os
import re
import json
import uuid
import base64
import sqlite3
import traceback
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from openai import OpenAI


# ──────────────────────────────────────────────────────────────────────────────
# App setup
# ──────────────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="LensVare Audit Intelligence",
    description="Natural language querying over audit Excel data — zero hallucination.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # tighten in production to your domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the frontend static files at /
app.mount("/app", StaticFiles(directory="frontend", html=True), name="frontend")


@app.get("/")
def root():
    return FileResponse("frontend/index.html")


# ──────────────────────────────────────────────────────────────────────────────
# In-memory session store
# Each session holds: conn (SQLite), schema (str), history (list), df_cols (list)
# ──────────────────────────────────────────────────────────────────────────────

sessions: dict[str, dict] = {}   # session_id → {conn, schema, history, col_names}

SESSION_TTL_REQUESTS = 200        # auto-expire after N queries (simple guard)


# ──────────────────────────────────────────────────────────────────────────────
# Providers catalogue
# ──────────────────────────────────────────────────────────────────────────────

PROVIDERS: dict[str, dict] = {
    "Kilo": {
        "base_url": "https://api.kilo.ai/api/gateway",
        "models": [
            "qwen3.6-Plus",
            "qwen3.6-Flash",
            "qwen3.6-Open-Source",
            "qwen3.5-Plus",
            "qwen3.5-Flash",
            "qwen3.5-Open-Source",
            "Kimi k2.6",
            "Kimi k2.5",
        ],
        "key_url": "https://kilo.ai/dashboard",
        "note": "Kilo AI gateway with Qwen 3.5/3.6 and Kimi K2 series models.",
    },
    "DashScope (Qwen)": {
        "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "models": [
            "qwen3-235b-a22b",
            "qwen3-30b-a3b",
            "qwen-max",
            "qwen-plus",
            "qwen-turbo",
            "qwen2.5-72b-instruct",
        ],
        "key_url": "https://dashscope.console.aliyun.com/apiKey",
        "note": "Qwen3-235B-A22B is the flagship model with extended thinking.",
    },
    "Moonshot (Kimi)": {
        "base_url": "https://api.moonshot.cn/v1",
        "models": [
            "kimi-k2",
            "kimi-k2-instruct",
            "moonshot-v1-128k",
            "moonshot-v1-32k",
            "moonshot-v1-8k",
        ],
        "key_url": "https://platform.moonshot.cn/console/api-keys",
        "note": "Kimi K2 and Moonshot models. moonshot-v1-128k for very large datasets.",
    },
    "OpenRouter (all models)": {
        "base_url": "https://openrouter.ai/api/v1",
        "models": [
            "qwen/qwen3.6-plus",
            "qwen/qwen3.6-flash",
            "qwen/qwen3.5-plus",
            "qwen/qwen3.5-flash",
            "moonshotai/kimi-k2",
            "moonshotai/kimi-k2-instruct",
            "qwen/qwen3-235b-a22b",
            "qwen/qwen3-32b",
            "meta-llama/llama-3.3-70b-instruct",
            "deepseek/deepseek-r1",
            "google/gemma-4-26b-a4b-it:free",
            "google/gemma-4-31b-it:free",
            "openai/gpt-oss-120b:free",
            "z-ai/glm-4.5-air:free",
            "qwen/qwen3-coder:free",
        ],
        "key_url": "https://openrouter.ai/keys",
        "note": "One key for every model. Best for comparing providers.",
    },
    "Together.ai": {
        "base_url": "https://api.together.xyz/v1",
        "models": [
            "Qwen/Qwen3-235B-A22B",
            "Qwen/Qwen2.5-72B-Instruct-Turbo",
            "meta-llama/Meta-Llama-3.1-70B-Instruct-Turbo",
            "deepseek-ai/DeepSeek-R1",
        ],
        "key_url": "https://api.together.ai/settings/api-keys",
        "note": "Low-latency inference for open-weight models.",
    },
}

# Env-var name for each provider's API key.
# If the env var is set, the server uses it and the frontend hides the key field.
PROVIDER_ENV_KEYS: dict[str, str] = {
    "Kilo":                     "KILO_API_KEY",
    "DashScope (Qwen)":         "DASHSCOPE_API_KEY",
    "Moonshot (Kimi)":          "MOONSHOT_API_KEY",
    "OpenRouter (all models)":  "OPENROUTER_API_KEY",
    "Together.ai":              "TOGETHER_API_KEY",
}


def _resolve_api_key(provider: str, client_key: str) -> str:
    """Return the server-side env key for *provider* if set, else the client-supplied key."""
    env_var = PROVIDER_ENV_KEYS.get(provider)
    if env_var:
        server_key = os.environ.get(env_var, "").strip()
        if server_key:
            return server_key
    return client_key


# ──────────────────────────────────────────────────────────────────────────────
# Data helpers (identical logic to app.py, now as pure functions)
# ──────────────────────────────────────────────────────────────────────────────

def _safe_col(name: str) -> str:
    s = re.sub(r"[^\w]", "_", name.strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return ("col_" + s) if s and s[0].isdigit() else (s or "unnamed")


def _safe_table(name: str) -> str:
    """Sanitize a sheet name to a valid SQLite table identifier."""
    s = re.sub(r"[^\w]", "_", str(name).strip())
    s = re.sub(r"_+", "_", s).strip("_").lower()
    return ("tbl_" + s) if (s and s[0].isdigit()) else (s or "sheet")


def load_excel_bytes(
    raw_bytes: bytes, filename: str
) -> tuple[dict[str, pd.DataFrame], sqlite3.Connection, dict[str, dict]]:
    """Load ALL sheets from an Excel file into one in-memory SQLite database.

    Returns
    -------
    sheets   : {table_name: DataFrame} — one entry per non-empty sheet
    conn     : in-memory SQLite connection with every sheet as its own table
    col_maps : {table_name: {original_col_name: sanitized_col_name}}
    """
    engine = "xlrd" if filename.endswith(".xls") else "openpyxl"
    all_sheets: dict = pd.read_excel(BytesIO(raw_bytes), engine=engine, sheet_name=None)

    conn = sqlite3.connect(":memory:", check_same_thread=False)
    sheets: dict[str, pd.DataFrame] = {}
    col_maps: dict[str, dict] = {}

    for sheet_name, raw in all_sheets.items():
        # Skip sheets that are entirely blank
        if raw.empty or raw.dropna(how="all").empty:
            continue

        # Sanitize sheet name → valid SQL table name; handle name collisions
        base = _safe_table(str(sheet_name))
        table_name = base
        n_dup = 2
        while table_name in sheets:
            table_name = f"{base}_{n_dup}"
            n_dup += 1

        # Sanitize column names within this sheet
        originals = raw.columns.tolist()
        seen: dict[str, int] = {}
        clean: list[str] = []
        for c in originals:
            base_col = _safe_col(str(c))
            if base_col in seen:
                seen[base_col] += 1
                clean.append(f"{base_col}_{seen[base_col]}")
            else:
                seen[base_col] = 0
                clean.append(base_col)

        raw = raw.copy()
        raw.columns = clean
        col_maps[table_name] = dict(zip(originals, clean))
        raw.to_sql(table_name, conn, if_exists="replace", index=False)
        sheets[table_name] = raw

    if not sheets:
        raise ValueError("Excel file contains no non-empty sheets.")

    return sheets, conn, col_maps


def build_schema(sheets: dict[str, pd.DataFrame]) -> str:
    """Build an LLM-friendly schema description for all loaded tables."""
    lines = ["=== SQLITE DATABASE SCHEMA ===", ""]

    for table_name, df in sheets.items():
        lines += [
            f"TABLE NAME : {table_name}",
            f"TOTAL ROWS : {len(df):,}",
            "",
            "COLUMNS  (use these EXACT names in every SQL query)",
            "─" * 60,
        ]

        for col in df.columns:
            s = df[col]
            non_null = int(s.notna().sum())
            lines.append(f"\n{col}")

            if pd.api.types.is_numeric_dtype(s.dtype):
                lines.append("  type     : numeric")
                if non_null:
                    lines.append(f"  range    : {s.min()} → {s.max()}")
                    lines.append(f"  mean     : {s.mean():.2f}")
                lines.append(f"  non-null : {non_null:,} / {len(df):,}")

            elif pd.api.types.is_datetime64_any_dtype(s.dtype):
                lines.append("  type     : datetime")
                if non_null:
                    lines.append(f"  range    : {s.min().date()} → {s.max().date()}")

            else:
                uniq = s.dropna().unique()
                n = len(uniq)
                lines.append(f"  type     : text  ({n} distinct values)")
                if n <= 40:
                    vals = " | ".join(sorted(str(v) for v in uniq))
                    lines.append(f"  values   : {vals}")
                else:
                    top = s.dropna().value_counts().head(8).index.tolist()
                    lines.append(f"  top-8    : {' | '.join(str(v) for v in top)}")
                lines.append(f"  non-null : {non_null:,} / {len(df):,}")

        lines += ["", "─" * 30, ""]

    lines += [
        "─" * 60,
        "SQL NOTES",
        "  • Partial text  : column LIKE '%value%'",
        "  • Case-insensitive : LOWER(column) LIKE LOWER('%value%')",
        "  • NULL check    : column IS NULL / IS NOT NULL",
        "  • Aggregation   : COUNT(*), SUM(col), AVG(col), GROUP BY col",
        "  • Sorting       : ORDER BY col DESC/ASC  LIMIT n",
        "  • Multi-table   : SELECT a.col, b.col FROM table_a a JOIN table_b b ON a.id = b.id",
        "=== END SCHEMA ===",
    ]
    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
# LLM pipeline
# ──────────────────────────────────────────────────────────────────────────────

def _strip_think(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


def _extract_sql(raw: str) -> str:
    s = re.sub(r"```[a-z]*\n?", "", raw).replace("```", "").strip()
    m = re.search(r"(SELECT\b.+)", s, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else s.strip()


def gen_sql(client: OpenAI, model: str, schema: str, question: str, history: list) -> str:
    system = f"""You are a precise SQLite query generator for an audit management system.

{schema}

STRICT OUTPUT RULES:
1. Return ONLY the raw SQL SELECT statement — nothing else.
2. No markdown, no explanation, no comments.
3. Use the exact column names listed in the schema.
4. For text matching: use LIKE or exact equality depending on context.
5. For "why" / "what causes" questions: SELECT all relevant columns including
   any root cause, department, risk rating, and description columns.
6. Never use INSERT, UPDATE, DELETE, DROP, CREATE, or any mutating statement.
7. If the question cannot be answered with SQL, write:
   SELECT 'Cannot answer this with SQL' AS message;"""

    msgs = [{"role": "system", "content": system}]
    for h in history[-3:]:
        msgs.append({"role": "user",      "content": h["q"]})
        msgs.append({"role": "assistant", "content": h["sql"]})
    msgs.append({"role": "user", "content": question})

    resp = client.chat.completions.create(
        model=model, messages=msgs, temperature=0.0, max_tokens=600,
    )
    return _extract_sql(_strip_think(resp.choices[0].message.content))


def run_sql(conn: sqlite3.Connection, sql: str) -> tuple[list[dict], str | None]:
    try:
        cur = conn.cursor()
        cur.execute(sql)
        if not cur.description:
            return [], None
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()], None
    except Exception as exc:
        return [], str(exc)


def gen_response(client: OpenAI, model: str, question: str, sql: str, results: list[dict]) -> str:
    n = len(results)
    if n == 0:
        data_str = "The query returned zero records."
    elif n <= 60:
        data_str = json.dumps(results, default=str, indent=2)
    else:
        data_str = (
            f"Total matching records: {n}\n\n"
            "First 50 records (representative sample):\n"
            + json.dumps(results[:50], default=str, indent=2)
            + f"\n\n[... {n - 50} additional records with similar structure]"
        )

    system = """You answer questions about audit data directly and concisely.

RULES:
• Use only facts from the query results — never invent numbers, names, or causes.
• For simple lookups (counts, names, values): give a direct one-sentence answer.
• For analytical questions (trends, why, patterns): give a brief analysis — 1 to 2 short paragraphs max.
• Never write generic boilerplate or pad the response.
• If zero records matched, say so and state what filter was applied."""

    user = (
        f"Question: {question}\n\n"
        f"SQL executed:\n{sql}\n\n"
        f"Data ({n} records):\n{data_str}\n\n"
        "Answer the question using only the data above. Be direct and brief."
    )

    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
        temperature=0.3,
        max_tokens=900,
    )
    return _strip_think(resp.choices[0].message.content)


def make_excel_bytes(data: dict[str, list[dict]]) -> bytes:
    """Build a styled Excel workbook from {sheet_name: [row_dict, ...]}.  Returns bytes."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, rows in data.items():
            df = pd.DataFrame(rows) if rows else pd.DataFrame()
            safe_name = sheet_name[:31]   # Excel sheet name limit
            df.to_excel(writer, sheet_name=safe_name, index=False)
            if not df.empty:
                ws = writer.sheets[safe_name]
                header_fill = PatternFill(start_color="1E2130", end_color="1E2130", fill_type="solid")
                header_font = Font(bold=True, color="C5CBF5")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for col_idx, col_cells in enumerate(ws.columns, 1):
                    max_len = max(
                        (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
                ws.freeze_panes = "A2"
    buf.seek(0)
    return buf.read()


def _extract_json_obj(text: str) -> dict:
    """Robustly extract the first top-level JSON object from LLM output.

    Handles: markdown fences, preamble text, trailing commentary, <think> tags.
    Raises json.JSONDecodeError if no valid JSON object found.
    """
    # Remove <think>...</think> blocks first
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
    # Remove ALL markdown code fences
    text = re.sub(r"```(?:json)?\s*", "", text)
    text = text.replace("```", "").strip()
    # Find the first '{' and walk to its matching '}'
    start = text.find("{")
    if start == -1:
        raise ValueError(f"No JSON object '{{' found in LLM response. Raw: {text[:300]!r}")
    depth = 0
    end = -1
    in_str = False
    escape = False
    for i, ch in enumerate(text[start:], start):
        if escape:
            escape = False
            continue
        if ch == "\\" and in_str:
            escape = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    if end == -1:
        raise ValueError(f"Unmatched braces in LLM response. Raw: {text[:300]!r}")
    fragment = text[start : end + 1]
    try:
        return json.loads(fragment)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON parse error: {e}. Fragment: {fragment[:300]!r}") from e


def _clean_val(val):
    """Convert a single cell value to a plain JSON-serialisable Python type."""
    if val is None:
        return None
    if hasattr(val, "item"):          # numpy scalar
        val = val.item()
    if hasattr(val, "isoformat"):     # Timestamp / datetime
        return val.isoformat()[:10]
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def gen_sample_rows(
    sheets: dict[str, pd.DataFrame],
    n_rows,                          # int for sampling, None for ALL rows
) -> dict[str, list[dict]]:
    """Generate rows per table.

    n_rows=None  → export every row as-is (complete file).
    n_rows=int   → draw *n_rows* random rows by sampling per-column value pools.
    No LLM call — instant, always succeeds.
    Returns {table_name: [{col: val, ...}, ...]}
    """
    import random

    result: dict[str, list[dict]] = {}
    for table, df in sheets.items():
        if n_rows is None:
            # Complete export — iterate all rows, clean types
            rows = [
                {col: _clean_val(row[col]) for col in df.columns}
                for _, row in df.iterrows()
            ]
        else:
            # Sampled export — build per-column pools, then draw n_rows
            pools: dict[str, list] = {}
            for col in df.columns:
                s = df[col].dropna()
                if len(s) == 0:
                    pools[col] = [None]
                elif pd.api.types.is_numeric_dtype(s.dtype):
                    pools[col] = s.tolist()
                elif pd.api.types.is_datetime64_any_dtype(s.dtype):
                    pools[col] = [str(v)[:10] for v in s.tolist()]
                else:
                    uniq = list(s.unique())
                    pools[col] = uniq if len(uniq) <= 200 else s.sample(200, replace=False).tolist()

            rows = []
            for _ in range(n_rows):
                row: dict = {}
                for col in df.columns:
                    row[col] = _clean_val(random.choice(pools[col]))
                rows.append(row)

        result[table] = rows

    total = sum(len(v) for v in result.values())
    print(f"[gen_sample_rows] {total} rows (n_rows={n_rows}) for tables: {list(result.keys())}", flush=True)
    return result


# ──────────────────────────────────────────────────────────────────────────────
# Request / Response models
# ──────────────────────────────────────────────────────────────────────────────

class AskRequest(BaseModel):
    session_id: str
    question: str
    provider: str
    model: str
    api_key: str


class AskResponse(BaseModel):
    answer: str
    sql: str
    row_count: int
    error: Optional[str] = None


class UploadResponse(BaseModel):
    session_id: str
    row_count: int
    col_count: int
    columns: list[str]
    schema_preview: str        # first 800 chars of the schema for display
    sheets: dict[str, list[str]]   # {table_name: [col1, col2, ...]} per sheet


# ──────────────────────────────────────────────────────────────────────────────
# API Routes
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "sessions_active": len(sessions)}


@app.get("/providers")
def list_providers():
    return {
        name: {
            "models": data["models"],
            "key_url": data["key_url"],
            "note": data["note"],
            # True when a server-side env key is configured for this provider
            "key_preset": bool(
                os.environ.get(PROVIDER_ENV_KEYS.get(name, ""), "").strip()
            ),
        }
        for name, data in PROVIDERS.items()
    }


@app.post("/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...)):
    """
    Accept an Excel file upload.
    Parse it into SQLite, build schema, return a session_id.
    The session_id is used for all subsequent /ask calls.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are accepted.")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:  # 50 MB cap
        raise HTTPException(status_code=413, detail="File too large. Maximum 50 MB.")

    try:
        sheets, conn, col_maps = load_excel_bytes(content, file.filename)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    schema = build_schema(sheets)
    session_id = str(uuid.uuid4())

    all_cols = [col for df in sheets.values() for col in df.columns]
    sessions[session_id] = {
        "conn": conn,
        "schema": schema,
        "history": [],
        "col_names": {t: list(df.columns) for t, df in sheets.items()},
        "sheets": {t: df for t, df in sheets.items()},
        "request_count": 0,
        "last_rows": None,
    }

    return UploadResponse(
        session_id=session_id,
        row_count=sum(len(df) for df in sheets.values()),
        col_count=len(all_cols),
        columns=all_cols,
        schema_preview=schema[:800],
        sheets={t: list(df.columns) for t, df in sheets.items()},
    )


@app.get("/sessions/{session_id}/schema")
def get_schema(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found or expired.")
    return {"schema": sessions[session_id]["schema"]}


@app.delete("/sessions/{session_id}")
def delete_session(session_id: str):
    if session_id in sessions:
        sessions[session_id]["conn"].close()
        del sessions[session_id]
    return {"deleted": True}


@app.get("/sessions/{session_id}/export")
def export_excel(session_id: str):
    """Download last query results as a styled .xlsx file."""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found.")
    rows = sessions[session_id].get("last_rows")
    if not rows:
        raise HTTPException(
            status_code=404,
            detail="No query results to export. Run a query first.",
        )
    xlsx_bytes = make_excel_bytes({"Query Results": rows})
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lensvare_export.xlsx"},
    )


@app.post("/ask", response_model=AskResponse)
def ask(req: AskRequest):
    """
    Main query endpoint.

    Flow:
      1. Look up SQLite connection + schema for session_id
      2. Generate SQL using the LLM (temperature=0)
      3. Execute SQL against the in-memory SQLite database (real data, no hallucination)
      4. If SQL fails, retry once with the error message fed back
      5. Convert result rows into professional prose (temperature=0.3)
      6. Return prose + sql + row_count to frontend
    """
    if req.session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found. Please re-upload your file.")

    sess = sessions[req.session_id]

    # Guard against dangling large sessions
    sess["request_count"] += 1
    if sess["request_count"] > SESSION_TTL_REQUESTS:
        sess["history"] = sess["history"][-10:]  # keep only recent history

    # Validate provider
    if req.provider not in PROVIDERS:
        raise HTTPException(status_code=400, detail=f"Unknown provider: {req.provider}")

    prov = PROVIDERS[req.provider]
    try:
        client = OpenAI(api_key=_resolve_api_key(req.provider, req.api_key), base_url=prov["base_url"])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not initialise client: {exc}")

    # ── Step 1: Generate SQL ──────────────────────────────────────────────────
    try:
        sql = gen_sql(client, req.model, sess["schema"], req.question, sess["history"])
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"LLM error generating SQL: {exc}")

    # ── Step 2: Execute SQL ───────────────────────────────────────────────────
    rows, err = run_sql(sess["conn"], sql)

    # Auto-retry once if the SQL had a syntax error
    if err:
        retry_prompt = (
            f"The SQL you generated caused this error:\n{err}\n\n"
            f"Faulty SQL:\n{sql}\n\n"
            f"Original question: {req.question}\n\n"
            "Return ONLY the corrected SQL SELECT statement."
        )
        try:
            sql = gen_sql(client, req.model, sess["schema"], retry_prompt, [])
            rows, err = run_sql(sess["conn"], sql)
        except Exception:
            pass

    if err:
        return AskResponse(
            answer=f"I was unable to retrieve an answer. The database query failed: {err}",
            sql=sql,
            row_count=0,
            error=err,
        )

    # ── Step 3: Generate prose response ──────────────────────────────────────
    try:
        answer = gen_response(client, req.model, req.question, sql, rows)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"LLM error generating response: {exc}")

    # Update history and cache last result for export
    sess["history"].append({"q": req.question, "sql": sql})
    sess["last_rows"] = rows

    return AskResponse(
        answer=answer,
        sql=sql,
        row_count=len(rows),
    )


@app.post("/ask/stream")
def ask_stream(req: AskRequest):
    """
    Streaming version of /ask.  Emits SSE events:
      {type: "meta",  sql: "...", row_count: N}  — after SQL executes
      {type: "token", text: "..."}               — one prose chunk
      {type: "done"}                             — stream complete
      {type: "error", message: "..."}            — on failure
    """
    if req.session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found. Please re-upload your file.")

    sess = sessions[req.session_id]
    sess["request_count"] += 1
    if sess["request_count"] > SESSION_TTL_REQUESTS:
        sess["history"] = sess["history"][-10:]

    if req.provider not in PROVIDERS:
        raise HTTPException(status_code=400, detail=f"Unknown provider: {req.provider}")

    prov = PROVIDERS[req.provider]
    try:
        client = OpenAI(api_key=_resolve_api_key(req.provider, req.api_key), base_url=prov["base_url"])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not initialise client: {exc}")

    def event_stream():
        # ── Detect keyword flags ──────────────────────────────────
        is_visualize = '@visualize' in req.question.lower()
        is_generate  = '@generate'  in req.question.lower()
        clean_q = re.sub(r'@(?:visualize|generate)\s*', '', req.question, flags=re.IGNORECASE).strip()

        # ── @generate: skip SQL, produce a sample Excel file ──────
        if is_generate:
            q_lower = clean_q.lower()

            # ── Row count: "complete / all / full / entire" → all rows; else parse number ──
            is_complete = bool(re.search(r'\b(complete|all|full|entire)\b', q_lower))
            n_match     = re.search(r'\b(\d+)\b', clean_q)
            if is_complete:
                n_rows = None                              # export every row
            elif n_match:
                n_rows = min(max(int(n_match.group(1)), 1), 10_000)
            else:
                n_rows = 5                                 # default sample

            # ── Optional column filter: fuzzy-match column names mentioned in the question ──
            filtered_sheets: dict[str, pd.DataFrame] = {}
            wanted_cols: list[str] = []
            for tname, df in sess["sheets"].items():
                wanted = [
                    col for col in df.columns
                    if col.lower().replace("_", " ") in q_lower
                    or col.lower() in q_lower
                ]
                filtered_sheets[tname] = df[wanted] if wanted else df
                wanted_cols.extend(wanted if wanted else df.columns.tolist())

            # ── Build human-readable label for the frontend message ──
            total_src_rows = sum(len(df) for df in sess["sheets"].values())
            if n_rows is None:
                row_label = f"complete file · {total_src_rows} rows"
            else:
                row_label = f"{n_rows} sample row{'s' if n_rows != 1 else ''}"

            all_orig_cols = [c for df in sess["sheets"].values() for c in df.columns]
            if sorted(wanted_cols) != sorted(all_orig_cols):
                col_label = " · columns: " + ", ".join(dict.fromkeys(wanted_cols))
            else:
                col_label = ""

            label = row_label + col_label

            try:
                sample     = gen_sample_rows(filtered_sheets, n_rows)
                xlsx_bytes = make_excel_bytes(sample)
                b64        = base64.b64encode(xlsx_bytes).decode()
                yield f"data: {json.dumps({'type': 'generate', 'filename': 'sample_lensvare.xlsx', 'data': b64, 'label': label})}\n\n"
            except Exception as exc:
                print(f"[gen_sample_rows ERROR] {type(exc).__name__}: {exc}", flush=True)
                yield f"data: {json.dumps({'type': 'error', 'message': f'Sample generation failed: {exc}'})}\n\n"
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
            return

        # ── Phase 1: Generate SQL (blocking, fast) ───────────────
        try:
            sql = gen_sql(client, req.model, sess["schema"], clean_q, sess["history"])
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'message': str(exc)})}\n\n"
            return

        # ── Phase 2: Execute SQL ─────────────────────────────────
        rows, err = run_sql(sess["conn"], sql)
        if err:
            retry_prompt = (
                f"The SQL you generated caused this error:\n{err}\n\n"
                f"Faulty SQL:\n{sql}\n\n"
                f"Original question: {clean_q}\n\n"
                "Return ONLY the corrected SQL SELECT statement."
            )
            try:
                sql = gen_sql(client, req.model, sess["schema"], retry_prompt, [])
                rows, err = run_sql(sess["conn"], sql)
            except Exception:
                pass

        if err:
            yield f"data: {json.dumps({'type': 'meta', 'sql': sql, 'row_count': 0})}\n\n"
            yield f"data: {json.dumps({'type': 'error', 'message': err})}\n\n"
            return

        # Cache results for export
        sess["last_rows"] = rows

        # Signal frontend: SQL ready, prose streaming starts now
        yield f"data: {json.dumps({'type': 'meta', 'sql': sql, 'row_count': len(rows)})}\n\n"

        # ── Phase 3: Chart or Prose ───────────────────────────────
        n = len(rows)
        if n == 0:
            data_str = "The query returned zero records."
        elif n <= 60:
            data_str = json.dumps(rows, default=str, indent=2)
        else:
            data_str = (
                f"Total matching records: {n}\n\n"
                "First 50 records (representative sample):\n"
                + json.dumps(rows[:50], default=str, indent=2)
                + f"\n\n[... {n - 50} additional records with similar structure]"
            )

        if is_visualize:
            # ── Chart generation (non-streaming) ─────────────────
            chart_system = (
                "You are a Chart.js v4 configuration generator.\n"
                "Output ONLY a valid Chart.js config JSON object — "
                "no prose, no markdown code fences, no explanation.\n\n"
                "RULES:\n"
                "- Use the actual data values from the query results.\n"
                "- Infer chart type from the request (pie, bar, line, doughnut, radar); default to bar.\n"
                "- For labels: use the first categorical/string column in the results.\n"
                "- For data values: use the first numeric column in the results.\n"
                "- backgroundColor array: [\"#4f6ef7\",\"#26c281\",\"#f5a623\",\"#e05252\","
                "\"#9b59b6\",\"#1abc9c\",\"#e67e22\",\"#3498db\"]\n"
                "- Set all text/tick/legend colors to \"#e8eaf6\" (dark theme).\n"
                "- For bar/line charts include scales with ticks.color and grid.color \"#2e3349\".\n"
                "- Output must be parseable by JSON.parse() with no surrounding text."
            )
            chart_user = (
                f"User request: {clean_q}\n\n"
                f"SQL executed:\n{sql}\n\n"
                f"Data ({n} records):\n{data_str}\n\n"
                "Output the Chart.js v4 config JSON object:"
            )
            try:
                resp = client.chat.completions.create(
                    model=req.model,
                    messages=[
                        {"role": "system", "content": chart_system},
                        {"role": "user",   "content": chart_user},
                    ],
                    temperature=0.1,
                    max_tokens=1500,
                )
                raw = _strip_think(resp.choices[0].message.content)
                raw = re.sub(r'^```(?:json)?\s*', '', raw.strip())
                raw = re.sub(r'\s*```$', '', raw.strip())
                config = json.loads(raw)
                yield f"data: {json.dumps({'type': 'chart', 'config': config})}\n\n"
            except json.JSONDecodeError:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Could not generate chart — try a more specific @visualize request.'})}\n\n"
                return
            except Exception as exc:
                yield f"data: {json.dumps({'type': 'error', 'message': str(exc)})}\n\n"
                return

        else:
            # ── Prose streaming ───────────────────────────────────
            system = """You answer questions about audit data directly and concisely.

RULES:
• Use only facts from the query results — never invent numbers, names, or causes.
• For simple lookups (counts, names, values): give a direct one-sentence answer.
• For analytical questions (trends, why, patterns): give a brief analysis — 1 to 2 short paragraphs max.
• Never write generic boilerplate or pad the response.
• If zero records matched, say so and state what filter was applied."""

            user = (
                f"Question: {clean_q}\n\n"
                f"SQL executed:\n{sql}\n\n"
                f"Data ({n} records):\n{data_str}\n\n"
                "Answer the question using only the data above. Be direct and brief."
            )

            in_think = False
            think_buf = ""
            try:
                stream = client.chat.completions.create(
                    model=req.model,
                    messages=[
                        {"role": "system", "content": system},
                        {"role": "user",   "content": user},
                    ],
                    temperature=0.3,
                    max_tokens=900,
                    stream=True,
                )
                for chunk in stream:
                    delta = (chunk.choices[0].delta.content or "") if chunk.choices else ""
                    if not delta:
                        continue

                    if in_think:
                        think_buf += delta
                        if "</think>" in think_buf:
                            in_think = False
                            after = think_buf.split("</think>", 1)[1]
                            think_buf = ""
                            if after:
                                yield f"data: {json.dumps({'type': 'token', 'text': after})}\n\n"
                    else:
                        if "<think>" in delta:
                            parts = delta.split("<think>", 1)
                            before, rest = parts[0], parts[1]
                            in_think = True
                            think_buf = rest
                            if before:
                                yield f"data: {json.dumps({'type': 'token', 'text': before})}\n\n"
                        else:
                            yield f"data: {json.dumps({'type': 'token', 'text': delta})}\n\n"

            except Exception as exc:
                yield f"data: {json.dumps({'type': 'error', 'message': str(exc)})}\n\n"
                return

        sess["history"].append({"q": clean_q, "sql": sql})
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
